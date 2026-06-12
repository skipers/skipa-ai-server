#!/usr/bin/env python3
from __future__ import annotations

import datetime as dt
import json
import sys
from pathlib import Path
from typing import Any


EXCEL_COLUMN_MAP = {
    "management_number": "관리번호",
    "draft_title": "발명의 명칭(가제)",
    "final_title": "발명의 명칭(최종)",
    "business_field": "관련사업 분야",
    "tech_field": "관련기술 분야",
    "related_products": "관련제품",
    "filing_country": "출원국",
    "is_joint_application": "공동출원여부",
    "joint_applicant": "공동출원인명",
    "status": "상태",
    "application_date": "출원일",
    "registration_date": "등록일",
    "application_number": "출원번호",
    "registration_number": "등록번호",
    "expiry_date": "예상 소멸일",
}


def normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def emit_parsed_json(root: Path) -> None:
    for path in sorted(root.glob("*/parsed.json"), key=lambda item: item.parent.name):
        payload = json.loads(path.read_text(encoding="utf-8"))
        record = {
            "parsed_json_key": path.as_posix(),
            "directory_name": path.parent.name,
            "payload": payload,
        }
        print(json.dumps(record, ensure_ascii=False, separators=(",", ":")))


def emit_excel_rows(path: Path) -> None:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl is required to read the patent Excel file") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return

    header_index = {header: index for index, header in enumerate(headers) if header}
    missing_headers = [header for header in EXCEL_COLUMN_MAP.values() if header not in header_index]
    if missing_headers:
        raise SystemExit(f"Missing Excel headers: {', '.join(missing_headers)}")

    for row in rows:
        if not row or not any(cell is not None for cell in row):
            continue

        record = {}
        for field_name, column_name in EXCEL_COLUMN_MAP.items():
            index = header_index[column_name]
            value = row[index] if index < len(row) else None
            record[field_name] = normalize_cell(value)

        if not record.get("application_number") and not record.get("registration_number"):
            continue

        print(json.dumps(record, ensure_ascii=False, separators=(",", ":")))


def main() -> None:
    if len(sys.argv) != 3 or sys.argv[1] not in {"parsed", "excel"}:
        raise SystemExit(
            "Usage: patent_seed_sources_to_jsonl.py parsed <parsed-root> | excel <xlsx-path>"
        )

    mode = sys.argv[1]
    path = Path(sys.argv[2])
    if mode == "parsed":
        emit_parsed_json(path)
    else:
        emit_excel_rows(path)


if __name__ == "__main__":
    main()
