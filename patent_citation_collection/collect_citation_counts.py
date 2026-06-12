#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - runtime dependency guard
    raise SystemExit("openpyxl is required. Install it with: pip install openpyxl") from exc

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover - optional dependency guard
    load_dotenv = None


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = PROJECT_ROOT / "특허리스트_등록_20260423기준 (1).xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "patent_citation_collection/output/patent_citation_counts.csv"
DEFAULT_CACHE_DIR = PROJECT_ROOT / "patent_citation_collection/cache"


def configure_imports() -> None:
    sys.path.extend(
        [
            str(PROJECT_ROOT / "eval_logic/legacy/src"),
            str(PROJECT_ROOT / "eval_logic/src"),
        ]
    )


def load_env() -> None:
    if load_dotenv:
        load_dotenv(PROJECT_ROOT / ".env")


@dataclass(frozen=True)
class PatentRow:
    registration_number: str
    application_number: str
    title: str


def clean_text(value: Any) -> str:
    return str(value or "").strip()


def load_patents_from_excel(path: Path) -> list[PatentRow]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return []

    header_index = {header: index for index, header in enumerate(headers) if header}
    required = ["등록번호", "출원번호", "발명의 명칭(최종)"]
    missing = [header for header in required if header not in header_index]
    if missing:
        raise ValueError(f"Excel file is missing required headers: {', '.join(missing)}")

    patents: list[PatentRow] = []
    seen: set[tuple[str, str]] = set()
    for row in rows:
        if not row or not any(cell is not None for cell in row):
            continue
        registration_number = clean_text(row[header_index["등록번호"]])
        application_number = clean_text(row[header_index["출원번호"]])
        title = clean_text(row[header_index["발명의 명칭(최종)"]])
        if not registration_number and not application_number:
            continue
        key = (registration_number, application_number)
        if key in seen:
            continue
        seen.add(key)
        patents.append(
            PatentRow(
                registration_number=registration_number,
                application_number=application_number,
                title=title,
            )
        )
    return patents


def cache_name(row: PatentRow) -> str:
    identifier = row.registration_number or row.application_number or row.title
    safe = "".join(ch if ch.isalnum() else "_" for ch in identifier).strip("_")
    return f"{safe or 'unknown'}.json"


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def collect_one(client: Any, row: PatentRow, cache_dir: Path, use_cache: bool) -> dict[str, str | int]:
    from crawling.kipris_api_client import count_citing_documents, sanitize_error

    now = datetime.now().isoformat(timespec="seconds")
    cache_path = cache_dir / cache_name(row)
    if use_cache and cache_path.exists():
        cached = json.loads(cache_path.read_text(encoding="utf-8"))
        return {
            "registration_number": row.registration_number,
            "citation_count": int(cached.get("citation_count") or 0),
            "application_number": row.application_number,
            "title": row.title,
            "source": cached.get("source", "cache"),
            "status": cached.get("status", "cached"),
            "error": cached.get("error", ""),
            "fetched_at": cached.get("fetched_at", now),
        }

    if not row.application_number:
        result = {
            "citation_count": 0,
            "source": "default",
            "status": "missing_application_number",
            "error": "application_number is required for KIPRIS CitingService",
            "fetched_at": now,
        }
    elif not client.enabled:
        result = {
            "citation_count": 0,
            "source": "default",
            "status": "missing_api_key",
            "error": "KIPRIS_API_KEY is not set",
            "fetched_at": now,
        }
    else:
        try:
            raw = client.citing_info(row.application_number)
            result = {
                "citation_count": count_citing_documents(raw),
                "source": "kipris_citing_service",
                "status": "ok",
                "error": "",
                "fetched_at": now,
                "raw": raw,
            }
        except Exception as exc:
            result = {
                "citation_count": 0,
                "source": "default",
                "status": "error",
                "error": sanitize_error(exc),
                "fetched_at": now,
            }

    write_json(
        cache_path,
        {
            "registration_number": row.registration_number,
            "application_number": row.application_number,
            "title": row.title,
            **result,
        },
    )
    return {
        "registration_number": row.registration_number,
        "citation_count": int(result.get("citation_count") or 0),
        "application_number": row.application_number,
        "title": row.title,
        "source": str(result.get("source") or ""),
        "status": str(result.get("status") or ""),
        "error": str(result.get("error") or ""),
        "fetched_at": str(result.get("fetched_at") or now),
    }


def write_csv(path: Path, rows: list[dict[str, str | int]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "registration_number",
        "citation_count",
        "application_number",
        "title",
        "source",
        "status",
        "error",
        "fetched_at",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def collect_citation_counts(
    excel_path: Path,
    output_path: Path,
    cache_dir: Path,
    limit: int | None,
    use_cache: bool,
    sleep_seconds: float,
) -> list[dict[str, str | int]]:
    configure_imports()
    load_env()
    from crawling.kipris_api_client import KiprisApiClient

    patents = load_patents_from_excel(excel_path)
    if limit is not None:
        patents = patents[:limit]

    client = KiprisApiClient(sleep_seconds=sleep_seconds)
    rows: list[dict[str, str | int]] = []
    for index, patent in enumerate(patents, 1):
        print(f"[{index}/{len(patents)}] {patent.registration_number or '-'} / {patent.application_number or '-'}")
        row = collect_one(client, patent, cache_dir, use_cache=use_cache)
        rows.append(row)
        print(f"  -> citation_count={row['citation_count']} status={row['status']}")

    write_csv(output_path, rows)
    print(f"\nSaved: {output_path} ({len(rows)} rows)")
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Collect patent forward citation counts into CSV.")
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL), help="Patent list .xlsx path")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output CSV path")
    parser.add_argument("--cache-dir", default=str(DEFAULT_CACHE_DIR), help="Raw/result cache directory")
    parser.add_argument("--limit", type=int, default=None, help="Collect only the first N rows")
    parser.add_argument("--no-cache", action="store_true", help="Ignore existing cache files")
    parser.add_argument("--sleep-seconds", type=float, default=0.2, help="Delay after each API call")
    args = parser.parse_args()

    collect_citation_counts(
        excel_path=Path(args.excel),
        output_path=Path(args.output),
        cache_dir=Path(args.cache_dir),
        limit=args.limit,
        use_cache=not args.no_cache,
        sleep_seconds=args.sleep_seconds,
    )


if __name__ == "__main__":
    main()
